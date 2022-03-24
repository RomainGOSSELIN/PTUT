from tkinter import Button
import streamlit as st
import os
import pandas as pd
import numpy as np
import xlsxwriter
import datetime
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb
from openpyxl import load_workbook

# Paramètres de la page
st.set_page_config(page_title = "STID ET APRES ?", page_icon = "🎓")
st.image("IMG/graduation_hat.png", width = 100)
st.title("STID ET APRES ?")

# Importation de la DB
df = pd.read_excel("formations.xlsx", header = 0)

# Fonction de conversion DF -> XLSX
def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine = 'xlsxwriter')
    df.to_excel(writer, index = False, sheet_name = 'Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)  
    writer.save()
    processed_data = output.getvalue()
    return processed_data

# Conversion df restreint 
formations = to_excel(df)

# Bouton de téléchargement de la db
st.download_button(label = "📊 Télécharger la base de données", data = formations, file_name = 'formations.xlsx')

# Titre - Critères
st.title("✔️ Critères", anchor = None)

# Critères de recherche
region = st.selectbox("Région 🗺️", df["Région"].unique())
df_restreint = df.loc[(df["Région"] == (region))]
departement = st.selectbox("Département 📍", df_restreint["Département"].unique())
diplome = st.selectbox("Diplôme délivré 🏆", df_restreint["Diplôme délivré"].unique())
statut = st.selectbox("Statut 💼", df_restreint["Statut"].unique())

# Titre - Résultat
st.title("🔍 Résultats", anchor = None)

# Sélection des lignes selon le respect des critères
df_restreint = df.loc[(df["Département"] == (departement)) & (df["Diplôme délivré"] == (diplome)) & (df["Statut"] == (statut))]

if len(df_restreint) != 0 & len(df_restreint) != 1:
    st.success("🥳 Hourra ! " + str(len(df_restreint)) + " formations correspondent à vos critères de recherche !")
elif len(df_restreint) == 1:
    st.success("🎉 Ouuuf ! " + str(len(df_restreint)) + " formation correspond à vos critères de recherche !")
elif len(df_restreint) == 0 :
    st.error("❌ Sapristi ! Aucune formation ne correspond à vos critères de recherche ! Essayez d'en changer !")

# Affichage du dataframe restreint
st.dataframe(df_restreint)

# Conversion df restreint 
liste_formations = to_excel(df_restreint)
# Bouton de téléchargement du résultat
st.download_button(label='✅ Télécharger le résultat', data = liste_formations, file_name = 'liste_formations.xlsx')

# Titre - Questionnaire
st.title("📋 Questionnaire", anchor = None)

with st.container():
    prenom = st.text_input("Indiquez votre prénom :")
    nom = st.text_input("Indiquez votre nom :")
    
    bac = st.radio("Quel Bac avez-vous fait ?",("Bac S (ou équivalent)", "Bac ES (ou équivalent)", "Bac L (ou équivalent)", "Bac STI2D", "Bac Professionnel", "Autre"))
    if bac == "Autre":
        bac = st.text_input("Précisez l'intitulé de votre Bac ?")
    mention = st.radio("Quelle a été votre mention ? ",("Très Bien", "Bien", "Assez Bien", "Sans mention"))

    choix = st.text_input("Pourquoi avoir choisi STID et spécialement celui d'Aurillac ?")
    integration = st.radio("Avez-vous intégré STID :", ("Directement après le Bac", "1 an après le Bac", "2 ans après le Bac", "3 ans ou plus après le Bac"))
    stage = st.text_input("Dans quelle entreprise avez-vous effectué votre stage de fin de DUT ?")
    satisfaction_STID = st.select_slider("A quel point avez-vous apprécié le DUT STID", options = ["Pas du tout", "Pas vraiment", "Moyennement", "Plutôt", "Énormément"])
        
    poursuite = st.radio("Que faites-vous actuellement ?", ("École d'ingénieur (en initial)", "École d'ingénieur (en alternance)", "Licence puis Master", "Licence professionelle", "Vie Active", "Autre"))
    if poursuite == "École d'ingénieur (en initial)" or poursuite == "École d'ingénieur (en alternance)" or poursuite == "Licence puis Master" or poursuite == "Licence professionelle":
        statut = st.text_input("Quel est l'intitulé de votre formation ?")
        ville = st.text_input("Dans quelle se déroule cette formation ?")
        choix = st.radio("Avez-vous été accepté dans la formation que vous préfériez ?", ("Oui", "Non"))
        satisfaction_formation = st.radio("Êtes-vous satisfait de votre formation actuelle ?", ("Oui", "Moyennement", "Non"))
        metier = st.text_input("Quel métier voudriez-vous faire après vos études ?")
    elif poursuite == "Vie Active":
        statut = st.text_input("Dans quelle entreprise travaillez-vous ?")
        ville = st.text_input("Dans quelle ville travaillez-vous ?")
        metier = st.text_input("Quel poste occupez-vous ?")
    elif poursuite == "Autre":
        statut = st.text_input("Expliquez-nous en quelques mots ce que vous faites / vos projets ?")

    recommandation = st.radio("Recommanderiez-vous STID à d'autres personnes ?", ("Oui", "Non"))
    if recommandation == "Non":
        pourquoi = st.text_input("Pourquoi ?")

    contact = st.text_input("Donnez-nous un moyen de vous recontacter pour de futurs forums d'anciens étudiants par exemple")
    autorisation = st.checkbox("Autorisez-vous l'utilisation et le traitement de vos données personnelles dans le but d'une analyse statistique ? (Les données personnelles telles que les notes et noms seront bien sûr anonymisées")

# Fonction ajout df à xlsx
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow = None, truncate_sheet = False, **to_excel_kwargs):
    if not os.path.isfile(filename):
        df.to_excel(filename, sheet_name = sheet_name, startrow = startrow if startrow is not None else 0, **to_excel_kwargs)
        return
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine = 'openpyxl', mode = 'a', if_sheet_exists = 'overlay')
    writer.book = load_workbook(filename)
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
        writer.book.create_sheet(sheet_name, idx)
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow = startrow, **to_excel_kwargs)
    writer.save()

# Bouton envoyer + réaction
envoyer = st.button('Envoyer')
if envoyer: 
    reponses = {"prenom" : prenom, "nom" : nom, "bac" : bac, "mention" : mention, "choix" : choix, "integration" : integration, "stage" : stage, "satisfaction_STID" : satisfaction_STID}
    df_reponses = pd.DataFrame(reponses, index = [datetime.datetime.now()])
    questionnaire = "questionnaire.xlsx"
    append_df_to_excel(questionnaire, df_reponses, header = False)
    st.success("💯 Vos réponses ont bien été enregistrées ! Merci de votre participation ! 🙏")